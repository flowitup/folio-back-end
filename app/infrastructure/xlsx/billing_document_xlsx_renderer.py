"""openpyxl renderer for BillingDocument entities.

Mirrors the source ANN ECO CONSTRUCTION xlsx layout exactly:

  Column grid (A through L, 12 cols):
    A=2.29, B=1.71, C=15.71, D=3.42, E=15.29, F=9.71,
    G=5.14, H=5.29, I=9.29, J=8.42, K=11.29, L=4.57
    (Description spans C..F via merged cells.)

  Row layout:
    row 1     A1: legal_name   — bold sz14 cyan (FF00A0DF), centered
    row 2-5   issuer header band (Siège / Email on left, SIREN/Tel/TVA/etc.)
    row 8     B="Réf Facture/Devis", D=document_number, H="À l'attention de :"
    rows 9-12 right column I = recipient name + address + tel
    row 12 B  "Objet/Opération"
    rows 13-14 C = project description / address
    row 17 B  "<City>, DD/MM/YYYY"
    row 19 B  "Madame, Monsieur,"
    rows 20-21 C = standard greeting paragraphs (wrap)
    row 23    items header — ORANGE BG (FFF18728), bold, centered, borders
              B=Libellé G=U H=Qté I=PU(HT)en€ J=Avancement K=Montant(HT)en€ L=TVA
    row 24    project name repeat row (bold dark grey, full-width-ish)
    rows 25+  section header rows (col C, bold, centered) interleaved with
              line-item rows (description merged C..F, plus G/H/I/J/K/L cells)
    row N+1   H="Total (HT)" K=SUM K-cells L="€"
    row N+2   H="TVA" K=Total*0.1 L="€"
    row N+3   H="Total (TTC)" K=K_HT+K_TVA L="€"
    row N+5   B="Veuillez agréer, …" (full-width)
    row N+7   B="COORDONNÉES BANCAIRES" — ORANGE BG, merged B..G, bold
    row N+8   B="Domiciliation" / D=value (merged D..L)
    row N+9   B="IBAN" / D=value
    row N+10  B="BIC" / D=value
    row N+12  A=Late-payment legal note — small font (sz=7), merged A..L

Returns raw bytes of an .xlsx file (Open Office XML).
"""

from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.domain.billing.document import BillingDocument
from app.domain.billing.enums import BillingDocumentKind


# ---------------------------------------------------------------------------
# Style helpers / constants
# ---------------------------------------------------------------------------

# Source-observed colors (ANN ECO CONSTRUCTION palette)
COLOR_TITLE_CYAN = "FF00A0DF"
COLOR_ORANGE_BG = "FFF18728"
COLOR_DARK_GREY = "FF222222"

# Column widths (Excel character units) — copied verbatim from the source.
COL_WIDTHS = {
    "A": 2.29,
    "B": 1.71,
    "C": 15.71,
    "D": 3.42,
    "E": 15.29,
    "F": 9.71,
    "G": 5.14,
    "H": 5.29,
    "I": 9.29,
    "J": 8.42,
    "K": 11.29,
    "L": 4.57,
}


def _font(size: int = 11, bold: bool = False, color: Optional[str] = None) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _fill(rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=rgb)


def _thin_box() -> Border:
    s = Side(style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h: str = "general", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Renderer
# ---------------------------------------------------------------------------


class OpenpyxlBillingDocumentXlsxRenderer:
    """Renders BillingDocument to xlsx bytes that visually match the
    ANN ECO CONSTRUCTION source format.

    Stateless. Instantiate once and call render() per document.
    """

    def render(self, doc: BillingDocument) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Devis" if doc.kind == BillingDocumentKind.DEVIS else "Facture"
        ws.sheet_view.showGridLines = False

        # Column widths (verbatim from source)
        for letter, width in COL_WIDTHS.items():
            ws.column_dimensions[letter].width = width

        kind_fr = "Devis" if doc.kind == BillingDocumentKind.DEVIS else "Facture"

        # ---- 1. Issuer header band (rows 1-5) ------------------------------
        # A1: legal name — full-width merged, bold sz14 cyan, centered.
        ws.cell(row=1, column=1, value=doc.issuer_legal_name)
        ws.cell(row=1, column=1).font = _font(14, bold=True, color=COLOR_TITLE_CYAN)
        ws.cell(row=1, column=1).alignment = _align("center", "center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

        # Row 2: address — merged C..L for room.
        ws.cell(row=2, column=3, value=f"Siège : {doc.issuer_address or ''}")
        ws.cell(row=2, column=3).font = _font(11)
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=12)

        # Row 4: SIRET on left (C..F), Tél on right (J..L) — placeholder ok.
        if doc.issuer_siret:
            ws.cell(row=4, column=3, value=f"SIRET {doc.issuer_siret}")
            ws.cell(row=4, column=3).font = _font(11)
            ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=8)

        # Row 5: TVA on right.
        if doc.issuer_tva_number:
            ws.cell(row=5, column=10, value=f"TVA  {doc.issuer_tva_number}")
            ws.cell(row=5, column=10).font = _font(11)
            ws.merge_cells(start_row=5, start_column=10, end_row=5, end_column=12)

        # ---- 2. Doc number row (row 8) -------------------------------------
        ws.cell(row=8, column=2, value=f"Réf {kind_fr}").font = _font(11, bold=True)
        ws.cell(row=8, column=4, value=doc.document_number).font = _font(11)
        ws.cell(row=8, column=8, value="À l'attention de :").alignment = _align("left")
        ws.cell(row=8, column=8).font = _font(11)

        # ---- 3. Recipient block (rows 9-12, col I merged I..L) -------------
        ws.cell(row=9, column=9, value=doc.recipient_name).font = _font(11, bold=True)
        ws.cell(row=9, column=9).alignment = _align("left", "bottom")
        ws.merge_cells(start_row=9, start_column=9, end_row=9, end_column=12)
        if doc.recipient_address:
            for offset, line in enumerate(doc.recipient_address.splitlines(), start=10):
                if offset > 12:
                    break
                ws.cell(row=offset, column=9, value=line).font = _font(11)
                ws.cell(row=offset, column=9).alignment = _align("left", "bottom")
                ws.merge_cells(start_row=offset, start_column=9, end_row=offset, end_column=12)

        # ---- 4. "Objet / Opération" (row 12 B + rows 13-14 C) ---------------
        ws.cell(row=12, column=2, value="Objet/Opération").font = _font(11, bold=True)
        # We don't have a dedicated project_title field. Use notes' first line
        # as project description if available, else fall back to "—".
        project_title = ""
        project_addr = ""
        if doc.notes:
            lines = [ln for ln in doc.notes.splitlines() if ln.strip()]
            if lines:
                project_title = lines[0]
            if len(lines) > 1:
                project_addr = lines[1]
        if not project_title:
            project_title = doc.recipient_name  # last-resort fallback
        ws.cell(row=13, column=3, value=project_title).font = _font(11)
        ws.merge_cells(start_row=13, start_column=3, end_row=13, end_column=8)
        if project_addr:
            ws.cell(row=14, column=3, value=project_addr).font = _font(11)
            ws.merge_cells(start_row=14, start_column=3, end_row=14, end_column=8)

        # ---- 5. Issue date (row 17 B) --------------------------------------
        # Format: "<City>, DD/MM/YYYY" — issuer_address line 2 (city) is best-effort.
        city = ""
        if doc.issuer_address:
            # Heuristic: take last comma-separated token before postal code
            addr_parts = [p.strip() for p in doc.issuer_address.split(",")]
            if addr_parts:
                # Try last part minus leading postcode
                last = addr_parts[-1]
                tokens = last.split()
                if tokens and tokens[0].isdigit():
                    city = " ".join(tokens[1:])
                else:
                    city = last
        date_str = doc.issue_date.strftime("%d/%m/%Y")
        line = f"{city}, {date_str}" if city else date_str
        ws.cell(row=17, column=2, value=line).font = _font(11)
        ws.merge_cells(start_row=17, start_column=2, end_row=17, end_column=8)

        # ---- 6. Greeting (rows 19-21) --------------------------------------
        ws.cell(row=19, column=2, value="Madame, Monsieur,").font = _font(11)
        intro_subject = "facture" if doc.kind == BillingDocumentKind.FACTURE else "devis"
        ws.cell(
            row=20,
            column=3,
            value=f"Veuillez trouver ci-après le {intro_subject} relatif à la mission citée en objet.",
        ).font = _font(11)
        ws.cell(row=20, column=3).alignment = _align(wrap=True)
        ws.merge_cells(start_row=20, start_column=3, end_row=20, end_column=12)
        ws.cell(
            row=21,
            column=3,
            value="Je reste à votre disposition pour toute précision ou complément d'information.",
        ).font = _font(11)
        ws.cell(row=21, column=3).alignment = _align(wrap=True)
        ws.merge_cells(start_row=21, start_column=3, end_row=21, end_column=12)
        ws.cell(row=19, column=2).font = _font(11)
        ws.merge_cells(start_row=19, start_column=2, end_row=19, end_column=12)

        # ---- 7. Items header (row 23) — ORANGE bg, bold, centered, borders -
        items_header_row = 23
        header_cells = [
            (2, "Libellé"),
            (7, "U"),
            (8, "Qté"),
            (9, "PU\n (HT) en €"),
            (10, "Avancement"),
            (11, "Montant (HT) en €"),
            (12, "TVA"),
        ]
        for col, label in header_cells:
            c = ws.cell(row=items_header_row, column=col, value=label)
            c.font = _font(11, bold=True)
            c.fill = _fill(COLOR_ORANGE_BG)
            c.alignment = _align("center", "center", wrap=True)
            c.border = _thin_box()
        # Description label spans B..F
        ws.merge_cells(start_row=items_header_row, start_column=2, end_row=items_header_row, end_column=6)

        # ---- 8. Project repeat row (row 24) — bold dark grey ---------------
        project_row = 24
        c = ws.cell(row=project_row, column=2, value=project_title)
        c.font = _font(11, bold=True, color=COLOR_DARK_GREY)
        c.alignment = _align("left", "center", wrap=True)
        c.border = _thin_box()
        ws.merge_cells(start_row=project_row, start_column=2, end_row=project_row, end_column=6)
        # Empty bordered cells for G-L
        for col in range(7, 13):
            ws.cell(row=project_row, column=col).border = _thin_box()

        # ---- 9. Items + section headers (row 25+) --------------------------
        row = items_header_row + 2  # row 25
        first_item_row = None
        last_item_row = None
        last_category: Optional[str] = None
        for item in doc.items:
            # Insert section header row when category changes
            if item.category and item.category != last_category:
                sh = ws.cell(row=row, column=3, value=item.category)
                sh.font = _font(11, bold=True)
                sh.alignment = _align("center", "center", wrap=True)
                sh.border = _thin_box()
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
                # Border the rest of the row (B + G..L) for visual continuity
                for col in [2] + list(range(7, 13)):
                    ws.cell(row=row, column=col).border = _thin_box()
                row += 1
                last_category = item.category
            elif not item.category:
                last_category = None

            # Item row
            if first_item_row is None:
                first_item_row = row
            last_item_row = row
            # Description (merged C..F)
            d = ws.cell(row=row, column=3, value=item.description)
            d.font = _font(11)
            d.alignment = _align("left", "center", wrap=True)
            d.border = _thin_box()
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            # Empty B (left margin) cell with border
            ws.cell(row=row, column=2).border = _thin_box()
            # G = unit (we don't model it — leave blank but bordered)
            ws.cell(row=row, column=7).border = _thin_box()
            ws.cell(row=row, column=7).alignment = _align("center", "center")
            # H = Qté
            ws.cell(row=row, column=8, value=float(item.quantity)).font = _font(11)
            ws.cell(row=row, column=8).alignment = _align("center", "center")
            ws.cell(row=row, column=8).border = _thin_box()
            # I = PU HT
            ws.cell(row=row, column=9, value=float(item.unit_price)).font = _font(11)
            ws.cell(row=row, column=9).alignment = _align("center", "center")
            ws.cell(row=row, column=9).number_format = "#,##0.00"
            ws.cell(row=row, column=9).border = _thin_box()
            # J = Avancement (default 100% — we don't model partial advancement)
            ws.cell(row=row, column=10, value=1).font = _font(11)
            ws.cell(row=row, column=10).alignment = _align("center", "center")
            ws.cell(row=row, column=10).number_format = "0.00%"
            ws.cell(row=row, column=10).border = _thin_box()
            # K = Montant HT — formula =I*J*H so Excel recomputes if user edits
            ws.cell(row=row, column=11, value=f"=I{row}*J{row}*H{row}").font = _font(11)
            ws.cell(row=row, column=11).alignment = _align("center", "center")
            ws.cell(row=row, column=11).number_format = "#,##0.00"
            ws.cell(row=row, column=11).border = _thin_box()
            # L = TVA (decimal: 0.1, 0.2)
            ws.cell(row=row, column=12, value=float(item.vat_rate) / 100).font = _font(11)
            ws.cell(row=row, column=12).alignment = _align("center", "center")
            ws.cell(row=row, column=12).number_format = "0.00%"
            ws.cell(row=row, column=12).border = _thin_box()
            row += 1

        # ---- 10. Totals (rows N+1 .. N+3) ---------------------------------
        if last_item_row is None:
            # No items — bail out before totals/footer to keep file valid.
            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()

        totals_row_ht = row
        totals_row_tva = row + 1
        totals_row_ttc = row + 2

        # Total HT
        ws.cell(row=totals_row_ht, column=8, value="Total (HT)").font = _font(11, bold=True)
        ws.cell(row=totals_row_ht, column=8).alignment = _align("center", "center")
        ws.cell(
            row=totals_row_ht,
            column=11,
            value=f"=SUM(K{first_item_row}:K{last_item_row})",
        ).font = _font(11, bold=True)
        ws.cell(row=totals_row_ht, column=11).number_format = "#,##0.00"
        ws.cell(row=totals_row_ht, column=12, value="€").font = _font(11, bold=True)
        ws.cell(row=totals_row_ht, column=12).alignment = _align("center", "center")

        # TVA — we use the doc's computed total_tva (which respects mixed VAT rates)
        # rather than =K_HT*0.1 (which only works for single-rate docs).
        ws.cell(row=totals_row_tva, column=8, value="TVA").font = _font(11, bold=True)
        ws.cell(row=totals_row_tva, column=8).alignment = _align("center", "center")
        ws.cell(row=totals_row_tva, column=11, value=float(doc.total_tva)).font = _font(11, bold=True)
        ws.cell(row=totals_row_tva, column=11).number_format = "#,##0.00"
        ws.cell(row=totals_row_tva, column=12, value="€").font = _font(11, bold=True)
        ws.cell(row=totals_row_tva, column=12).alignment = _align("center", "center")

        # Total TTC
        ws.cell(row=totals_row_ttc, column=8, value="Total (TTC)").font = _font(11, bold=True)
        ws.cell(row=totals_row_ttc, column=8).alignment = _align("center", "center")
        ws.cell(
            row=totals_row_ttc,
            column=11,
            value=f"=K{totals_row_ht}+K{totals_row_tva}",
        ).font = _font(11, bold=True)
        ws.cell(row=totals_row_ttc, column=11).number_format = "#,##0.00"
        ws.cell(row=totals_row_ttc, column=12, value="€").font = _font(11, bold=True)
        ws.cell(row=totals_row_ttc, column=12).alignment = _align("center", "center")

        # ---- 11. Closing greeting (row N+5) -------------------------------
        closing_row = totals_row_ttc + 2
        ws.cell(
            row=closing_row,
            column=2,
            value="Veuillez agréer, Madame, Monsieur, l'expression de nos salutations distinguées.",
        ).font = _font(11)
        ws.merge_cells(start_row=closing_row, start_column=2, end_row=closing_row, end_column=12)

        # ---- 12. Bank coords block (rows N+7 .. N+10) ---------------------
        if doc.issuer_iban or doc.issuer_bic:
            bank_title_row = closing_row + 2
            c = ws.cell(row=bank_title_row, column=2, value="COORDONNÉES BANCAIRES")
            c.font = _font(11, bold=True)
            c.fill = _fill(COLOR_ORANGE_BG)
            c.alignment = _align("center", "center")
            ws.merge_cells(start_row=bank_title_row, start_column=2, end_row=bank_title_row, end_column=7)
            offset = 1
            if doc.issuer_iban:
                r = bank_title_row + offset
                ws.cell(row=r, column=2, value="IBAN").font = _font(11)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                ws.cell(row=r, column=4, value=doc.issuer_iban).font = _font(11)
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
                offset += 1
            if doc.issuer_bic:
                r = bank_title_row + offset
                ws.cell(row=r, column=2, value="BIC").font = _font(11)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                ws.cell(row=r, column=4, value=doc.issuer_bic).font = _font(11)
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
                offset += 1
            footer_anchor = bank_title_row + offset + 1
        else:
            footer_anchor = closing_row + 2

        # ---- 13. Late-payment legal note (factures only) ------------------
        if doc.kind == BillingDocumentKind.FACTURE:
            note = (
                "Indemnité forfaitaire de retard de paiement: 40€ "
                "(conformément à l'article 121-II de la loi n° 2012-387 du 22 Mars 2012 "
                "et au décret n° 2012-1115 du 2 Oct. 2012)"
            )
            ws.cell(row=footer_anchor, column=1, value=note).font = _font(7)
            ws.cell(row=footer_anchor, column=1).alignment = _align("left", "top", wrap=True)
            ws.merge_cells(start_row=footer_anchor, start_column=1, end_row=footer_anchor, end_column=12)

        # Serialise
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
