"""Pure-python xlsx builder for labor export.

build_xlsx(context: ExportContext, buckets: list[MonthBucket]) -> bytes

Sheet layout
------------
Sheet 1 "Summary"   — aggregated per-worker rows across all months + footer totals
Sheets 2..N+1 "MMM YYYY" — per-worker monthly summary (top) + daily detail (bottom)

Currency rule (LOCKED)
-----------------------
All currency cells carry RAW float values so Excel can sort/sum them.
The cell.number_format is set to EUR_FR_FORMAT which renders the value as
  200,00 € in fr-FR locales — matching FE Intl.NumberFormat("fr-FR", {currency:"EUR"}).
Never write pre-formatted strings to currency cells.

"Total (priced + bonus)" column (reviewer HIGH-3 carry-forward)
---------------------------------------------------------------
The last column of every summary table is explicitly labelled
"Total (priced + bonus)" — never just "Total" — to avoid the
inflation-ambiguity flagged in the original code review.

No aggregated misleading Total: Priced cost + Bonus cost are
separate columns so readers can distinguish the two cost components.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.labor.export.models import ExportContext, MonthBucket

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EUR_FR_FORMAT = "_-* #,##0.00\\ [$€-fr-FR]_-;" "-* #,##0.00\\ [$€-fr-FR]_-;" '_-* "-"?? [$€-fr-FR]_-;' "_-@_-"

_SUMMARY_HEADERS = [
    "Worker",
    "Days",
    "Banked hrs",
    "Bonus full",
    "Bonus half",
    "Priced cost",
    "Bonus cost",
    "Total (priced + bonus)",
]

_DETAIL_HEADERS = [
    "Date",
    "Worker",
    "Shift",
    "Supplement hrs",
    "Override",
    "Effective cost",
    "Note",
]

# Column widths per section header label
_SUMMARY_COL_WIDTHS = [30, 12, 12, 12, 12, 15, 15, 22]
_DETAIL_COL_WIDTHS = [12, 30, 12, 14, 15, 15, 30]

# Openpyxl column letters for summary / detail
_SUMMARY_COLS = [get_column_letter(i + 1) for i in range(len(_SUMMARY_HEADERS))]
_DETAIL_COLS = [get_column_letter(i + 1) for i in range(len(_DETAIL_HEADERS))]

# Currency column indices (0-based) in summary table: Priced cost=5, Bonus cost=6, Total=7
_SUMMARY_CURRENCY_COLS = {5, 6, 7}
# Currency column index in detail table: Effective cost=5
_DETAIL_CURRENCY_COLS = {5}


# ---------------------------------------------------------------------------
# Internal aggregation dataclass
# ---------------------------------------------------------------------------


@dataclass
class _AggRow:
    worker_id: str
    worker_name: str
    days_worked: int
    banked_hours: int
    bonus_full_days: int
    bonus_half_days: int
    priced_cost: float  # total_cost - bonus_cost
    bonus_cost: float
    total_cost: float  # priced_cost + bonus_cost


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def _bold_font(size: int = 11) -> Font:
    return Font(bold=True, size=size)


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _thick_top_border() -> Border:
    thick = Side(style="medium")
    return Border(top=thick)


def _header_fill() -> PatternFill:
    return PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _apply_border_row(ws: Worksheet, row: int, n_cols: int, border: Border) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).border = border


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------


def _aggregate_across_buckets(buckets: List[MonthBucket]) -> List[_AggRow]:
    """Sum per-worker fields across all buckets, keyed by worker_id.

    Worker names may differ across buckets only in pathological cases; we use
    the name from the FIRST bucket encounter to stay deterministic.
    Worker rows are sorted by worker_name (Unicode-stable, no locale dependency).
    """
    agg: dict[str, _AggRow] = {}
    for bucket in buckets:
        for row in bucket.summary.rows:
            priced = row.total_cost - row.bonus_cost
            if row.worker_id not in agg:
                agg[row.worker_id] = _AggRow(
                    worker_id=row.worker_id,
                    worker_name=row.worker_name,
                    days_worked=row.days_worked,
                    banked_hours=row.banked_hours,
                    bonus_full_days=row.bonus_full_days,
                    bonus_half_days=row.bonus_half_days,
                    priced_cost=priced,
                    bonus_cost=row.bonus_cost,
                    total_cost=row.total_cost,
                )
            else:
                a = agg[row.worker_id]
                a.days_worked += row.days_worked
                a.banked_hours += row.banked_hours
                a.bonus_full_days += row.bonus_full_days
                a.bonus_half_days += row.bonus_half_days
                a.priced_cost += priced
                a.bonus_cost += row.bonus_cost
                a.total_cost += row.total_cost

    return sorted(agg.values(), key=lambda r: r.worker_name)


def _month_agg_rows(bucket: MonthBucket) -> List[_AggRow]:
    """Per-worker rows for a single bucket (no cross-bucket summing needed)."""
    rows: List[_AggRow] = []
    for r in bucket.summary.rows:
        priced = r.total_cost - r.bonus_cost
        rows.append(
            _AggRow(
                worker_id=r.worker_id,
                worker_name=r.worker_name,
                days_worked=r.days_worked,
                banked_hours=r.banked_hours,
                bonus_full_days=r.bonus_full_days,
                bonus_half_days=r.bonus_half_days,
                priced_cost=priced,
                bonus_cost=r.bonus_cost,
                total_cost=r.total_cost,
            )
        )
    return sorted(rows, key=lambda r: r.worker_name)


# ---------------------------------------------------------------------------
# Sheet-writing helpers
# ---------------------------------------------------------------------------


def _write_header_block(ws: Worksheet, context: ExportContext, month_label: str | None = None) -> int:
    """Write rows 1-4 (project header). Returns next_row (5)."""
    from_label = context.range.from_month.strftime("%b %Y")
    to_label = context.range.to_month.strftime("%b %Y")

    ws["A1"] = "Folio · Labor Export"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Project: {context.project_name}"

    if month_label:
        ws["A3"] = f"Month: {month_label}"
    else:
        ws["A3"] = f"Range: {from_label} → {to_label}"

    ws["A4"] = f"Generated: {context.generated_at.isoformat()} " f"by {context.generated_by_email}"

    # Row 5 blank
    return 5


def _write_summary_table(
    ws: Worksheet,
    start_row: int,
    agg_rows: List[_AggRow],
) -> int:
    """Write summary header + per-worker rows + footer totals.

    Returns the row number AFTER the footer row.
    """
    thin = _thin_border()
    thick = _thick_top_border()
    fill = _header_fill()

    # Header row
    hdr_row = start_row
    for i, label in enumerate(_SUMMARY_HEADERS):
        cell = ws.cell(row=hdr_row, column=i + 1, value=label)
        cell.font = _bold_font()
        cell.border = thin
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Per-worker rows
    data_start = hdr_row + 1
    for row_idx, agg in enumerate(agg_rows):
        r = data_start + row_idx
        values = [
            agg.worker_name,
            agg.days_worked,
            agg.banked_hours,
            agg.bonus_full_days,
            agg.bonus_half_days,
            agg.priced_cost,
            agg.bonus_cost,
            agg.total_cost,
        ]
        for col_idx, val in enumerate(values):
            cell = ws.cell(row=r, column=col_idx + 1, value=val)
            cell.border = thin
            if col_idx in _SUMMARY_CURRENCY_COLS:
                cell.number_format = EUR_FR_FORMAT
                cell.value = float(val)  # ensure float, not Decimal
            elif col_idx == 0:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Footer totals row
    footer_row = data_start + len(agg_rows)
    totals = [
        "TOTAL",
        sum(r.days_worked for r in agg_rows),
        sum(r.banked_hours for r in agg_rows),
        sum(r.bonus_full_days for r in agg_rows),
        sum(r.bonus_half_days for r in agg_rows),
        sum(r.priced_cost for r in agg_rows),
        sum(r.bonus_cost for r in agg_rows),
        sum(r.total_cost for r in agg_rows),
    ]
    for col_idx, val in enumerate(totals):
        cell = ws.cell(row=footer_row, column=col_idx + 1, value=val)
        cell.font = _bold_font()
        cell.border = thick
        if col_idx in _SUMMARY_CURRENCY_COLS:
            cell.number_format = EUR_FR_FORMAT
            cell.value = float(val)

    return footer_row + 1  # next available row


def _write_daily_detail(
    ws: Worksheet,
    start_row: int,
    entries: list,
) -> int:
    """Write 'Daily detail' section header + table. Returns next row after last data row."""
    thin = _thin_border()
    fill = _header_fill()

    # Section label
    section_cell = ws.cell(row=start_row, column=1, value="Daily detail")
    section_cell.font = _bold_font()

    # Table header row
    hdr_row = start_row + 1
    for i, label in enumerate(_DETAIL_HEADERS):
        cell = ws.cell(row=hdr_row, column=i + 1, value=label)
        cell.font = _bold_font()
        cell.border = thin
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Sort entries: date asc, worker_name asc
    sorted_entries = sorted(
        entries,
        key=lambda e: (e.date, e.worker_name),
    )

    data_row = hdr_row + 1
    for entry in sorted_entries:
        override_val = entry.amount_override if entry.amount_override is not None else ""
        values = [
            entry.date,
            entry.worker_name,
            entry.shift_type or "",
            entry.supplement_hours,
            override_val,
            entry.effective_cost,
            entry.note or "",
        ]
        for col_idx, val in enumerate(values):
            cell = ws.cell(row=data_row, column=col_idx + 1, value=val)
            cell.border = thin
            if col_idx in _DETAIL_CURRENCY_COLS:
                cell.number_format = EUR_FR_FORMAT
                cell.value = float(val) if val != "" else 0.0
            elif col_idx == 4 and val != "":  # Override column (currency if present)
                cell.number_format = EUR_FR_FORMAT
                cell.value = float(val)
        data_row += 1

    return data_row


def _set_summary_col_widths(ws: Worksheet) -> None:
    for i, width in enumerate(_SUMMARY_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def _set_detail_col_widths(ws: Worksheet) -> None:
    for i, width in enumerate(_DETAIL_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def _sanitize_sheet_name(name: str, fallback: str) -> str:
    """Return a valid Excel sheet name (≤31 chars, no forbidden chars).

    Excel forbids: [ ] : * ? / \\
    If the sanitized name is empty, fall back to the first 8 chars of fallback.
    """
    import re

    sanitized = re.sub(r"[\[\]:*?/\\]", "", name).strip()
    if not sanitized:
        sanitized = fallback[:8] or "Worker"
    return sanitized[:31]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def _build_xlsx_single_worker(
    context: ExportContext,
    buckets: List[MonthBucket],
    wb: openpyxl.Workbook,
) -> None:
    """Render a single-worker export into wb: one sheet with header + per-month tables.

    Sheet name: sanitized worker name (≤31 chars, Excel-safe).
    Layout per bucket (month):
      - Section label "MMM YYYY"
      - Monthly summary table (single worker row)
      - Blank separator
      - Daily detail table
    Empty-range: single "No labor entries…" message cell, no tables.
    """
    worker_name = context.worker_name or "Worker"
    sheet_title = _sanitize_sheet_name(worker_name, worker_name[:8])

    ws = wb.active
    ws.title = sheet_title

    next_row = _write_header_block(ws, context, month_label=None)

    # Add worker-specific sub-header (row 5 used by blank; write into row 5)
    rate = context.worker_daily_rate
    rate_str = str(rate) if rate is not None else "—"
    ws.cell(row=next_row, column=1, value=f"Worker: {worker_name}    Rate: {rate_str}/day").font = Font(italic=True)
    next_row += 1  # advance past worker sub-header (row 6 is now available for data)

    all_empty = all(not bucket.summary.rows and not bucket.daily_entries for bucket in buckets) if buckets else True

    if all_empty:
        from_label = context.range.from_month.strftime("%b %Y")
        to_label = context.range.to_month.strftime("%b %Y")
        ws.cell(row=next_row, column=1, value=f"No labor entries in range {from_label} → {to_label}").font = Font(
            italic=True
        )
        _set_summary_col_widths(ws)
        _set_detail_col_widths(ws)
        return

    for bucket in sorted(buckets, key=lambda b: b.month):
        month_label = bucket.month.strftime("%b %Y")

        # Month section label
        label_cell = ws.cell(row=next_row, column=1, value=month_label)
        label_cell.font = Font(bold=True, size=12)
        next_row += 1

        # Per-worker monthly summary table (may be empty if worker had no entries this month)
        month_rows = _month_agg_rows(bucket)
        if month_rows:
            next_row = _write_summary_table(ws, start_row=next_row, agg_rows=month_rows)
        else:
            ws.cell(row=next_row, column=1, value="No entries this month").font = Font(italic=True)
            next_row += 1

        _set_summary_col_widths(ws)

        # Blank separator
        next_row += 1

        # Daily detail for this month
        next_row = _write_daily_detail(ws, start_row=next_row, entries=bucket.daily_entries)
        _set_detail_col_widths(ws)

        # Extra blank row between months
        next_row += 1


def build_xlsx(context: ExportContext, buckets: List[MonthBucket]) -> bytes:
    """Generate an xlsx workbook and return its raw bytes.

    Single-worker mode (context.worker_name is set):
    - ONE sheet named after the worker
    - Layout: header block, worker sub-header, then per-month: summary table + daily detail

    Project-wide mode (context.worker_name is None):
    - Sheet 1 "Summary": aggregated per-worker across all months
    - Sheets 2..N+1 "MMM YYYY": per-worker monthly summary + daily detail

    Empty-range case: all buckets have empty summary.rows AND empty daily_entries.
    → Only the relevant sheet is written with a single "No labor entries" message cell.

    Currency cells store raw float values; number_format renders as fr-FR EUR in Excel.
    """
    wb = openpyxl.Workbook()

    # --- Single-worker mode ---
    if context.worker_name is not None:
        _build_xlsx_single_worker(context, buckets, wb)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # --- Project-wide mode ---

    # Check empty-range case
    all_empty = all(not bucket.summary.rows and not bucket.daily_entries for bucket in buckets) if buckets else True

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    _write_header_block(ws_summary, context, month_label=None)

    if all_empty:
        from_label = context.range.from_month.strftime("%b %Y")
        to_label = context.range.to_month.strftime("%b %Y")
        ws_summary["A6"] = f"No labor entries in range {from_label} → {to_label}"
        ws_summary["A6"].font = Font(italic=True)
        _set_summary_col_widths(ws_summary)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    agg_rows = _aggregate_across_buckets(buckets)
    _write_summary_table(ws_summary, start_row=6, agg_rows=agg_rows)
    _set_summary_col_widths(ws_summary)

    # --- Sheets 2..N+1: per month ---
    for bucket in sorted(buckets, key=lambda b: b.month):
        month_label = bucket.month.strftime("%b %Y")
        ws = wb.create_sheet(title=month_label)

        _write_header_block(ws, context, month_label=month_label)

        # Per-worker monthly summary (top section)
        month_rows = _month_agg_rows(bucket)
        next_month_row = _write_summary_table(ws, start_row=6, agg_rows=month_rows)
        _set_summary_col_widths(ws)

        # Blank row separator
        blank_row = next_month_row
        next_month_row = blank_row + 1

        # Daily detail section
        _write_daily_detail(ws, start_row=next_month_row, entries=bucket.daily_entries)
        _set_detail_col_widths(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
