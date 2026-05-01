"""Pure-python xlsx builder for invoice export.

build_xlsx(context: InvoiceExportContext, bundle: InvoiceBundle) -> bytes

Sheet layout
------------
Sheet 1 "Summary"   — KPI table, subtotals by type, full invoice list + grand total
Sheets 2..k         — one sheet per InvoiceType that has at least one invoice

Currency rule (LOCKED)
-----------------------
All currency cells carry RAW float values so Excel can sort/sum them.
The cell.number_format is set to EUR_FR_FORMAT (imported from labor xlsx_builder —
do NOT redefine it) which renders the value as 200,00 € in fr-FR locales.
Never write pre-formatted strings to currency cells.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Reuse the canonical EUR_FR_FORMAT constant — do not redefine.
from app.domain.labor.export.xlsx_builder import EUR_FR_FORMAT  # noqa: F401

from app.domain.invoice.export.format import TYPE_LABEL_EN
from app.domain.invoice.export.models import InvoiceBundle, InvoiceExportContext, TypeSubtotal
from app.domain.entities.invoice import Invoice, InvoiceType

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

_INVOICE_HEADERS = ["#", "Date", "Type", "Recipient", "Items", "Total"]
_INVOICE_COL_WIDTHS = [6, 14, 12, 30, 8, 18]

_TYPE_SHEET_HEADERS = ["#", "Date", "Recipient", "Items", "Total"]
_TYPE_SHEET_COL_WIDTHS = [6, 14, 30, 8, 18]

_SUBTYPE_HEADERS = ["Type", "Invoice count", "Total"]
_SUBTYPE_COL_WIDTHS = [16, 16, 18]

_KPI_COL_WIDTHS = [22, 30]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def _bold_font(size: int = 11) -> Font:
    return Font(bold=True, size=size)


def _italic_font(size: int = 9) -> Font:
    return Font(italic=True, size=size)


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _thick_top_border() -> Border:
    thick = Side(style="medium")
    return Border(top=thick)


def _header_fill() -> PatternFill:
    return PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _grand_total_fill() -> PatternFill:
    return PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def _set_col_widths(ws: Worksheet, widths: List[int]) -> None:
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _merge_and_write(ws: Worksheet, row: int, start_col: int, end_col: int, value: str, font: Font) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Invoice sort key
# ---------------------------------------------------------------------------


def _invoice_sort_key(inv: Invoice):
    return (inv.issue_date, inv.type.value, inv.invoice_number)


# ---------------------------------------------------------------------------
# Summary sheet helpers
# ---------------------------------------------------------------------------


def _write_summary_header_band(ws: Worksheet, context: InvoiceExportContext) -> int:
    """Row 1: title band. Row 2: meta line. Row 3: blank. Returns next_row (4)."""
    from_label = context.range.from_month.strftime("%Y-%m")
    to_label = context.range.to_month.strftime("%Y-%m")

    # Row 1 — title
    _merge_and_write(
        ws,
        row=1,
        start_col=1,
        end_col=6,
        value=f"INVOICE EXPORT — {context.project_name}",
        font=Font(bold=True, size=14),
    )

    # Row 2 — meta
    meta = (
        f"Range: {from_label} to {to_label}"
        f" · Generated {context.generated_at.strftime('%Y-%m-%dT%H:%M')}"
        f" by {context.generated_by_email}"
    )
    cell2 = ws.cell(row=2, column=1, value=meta)
    cell2.font = Font(italic=True, size=9)

    # Row 3 blank
    return 4


def _write_kpi_table(ws: Worksheet, start_row: int, context: InvoiceExportContext, bundle: InvoiceBundle) -> int:
    """Write 2-col KPI table (4 rows). Returns next_row after blank spacer."""
    from_label = context.range.from_month.strftime("%Y-%m")
    to_label = context.range.to_month.strftime("%Y-%m")

    kpi_rows = [
        ("Total invoices", bundle.invoice_count),
        ("Grand total", bundle.grand_total),
        ("From month", from_label),
        ("To month", to_label),
    ]

    thin = _thin_border()
    fill = _header_fill()

    for i, (label, value) in enumerate(kpi_rows):
        r = start_row + i
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = _bold_font(10)
        label_cell.fill = fill
        label_cell.border = thin

        if i == 1:  # Grand total row — write float once, apply currency format
            cell_value = float(bundle.grand_total)
            val_cell = ws.cell(row=r, column=2, value=cell_value)
            val_cell.number_format = EUR_FR_FORMAT
        else:
            val_cell = ws.cell(row=r, column=2, value=value)
            val_cell.alignment = Alignment(horizontal="left")
        val_cell.border = thin

    ws.column_dimensions["A"].width = _KPI_COL_WIDTHS[0]
    ws.column_dimensions["B"].width = _KPI_COL_WIDTHS[1]

    # Return row 9 (start_row=4, 4 kpi rows = rows 4-7, row 8 blank => row 9)
    return start_row + 4 + 1  # +4 data rows +1 blank


def _write_subtotals_section(ws: Worksheet, start_row: int, subtotals: List[TypeSubtotal]) -> int:
    """Write 'Subtotals by type' heading + header row + data rows. Returns next_row."""
    thin = _thin_border()
    fill = _header_fill()

    # Heading
    heading_cell = ws.cell(row=start_row, column=1, value="Subtotals by type")
    heading_cell.font = _bold_font(11)

    # Header row
    hdr_row = start_row + 1
    for i, label in enumerate(_SUBTYPE_HEADERS):
        cell = ws.cell(row=hdr_row, column=i + 1, value=label)
        cell.font = _bold_font()
        cell.border = thin
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data_row = hdr_row + 1
    for sub in subtotals:
        type_label = TYPE_LABEL_EN.get(sub.type.value, sub.type.value.title())
        cells_data = [type_label, sub.invoice_count, sub.total_amount]
        for col_idx, val in enumerate(cells_data):
            cell = ws.cell(row=data_row, column=col_idx + 1, value=val)
            cell.border = thin
            if col_idx == 2:  # Total column — currency
                cell.value = float(sub.total_amount)
                cell.number_format = EUR_FR_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
        data_row += 1

    return data_row + 1  # blank spacer


def _write_invoices_section(ws: Worksheet, start_row: int, invoices: List[Invoice]) -> int:
    """Write 'Invoices' heading + header + data rows + GRAND TOTAL band. Returns next_row."""
    thin = _thin_border()
    thick = _thick_top_border()
    fill = _header_fill()
    grand_fill = _grand_total_fill()

    # Section heading
    heading_cell = ws.cell(row=start_row, column=1, value="Invoices")
    heading_cell.font = _bold_font(11)

    # Header row
    hdr_row = start_row + 1
    for i, label in enumerate(_INVOICE_HEADERS):
        cell = ws.cell(row=hdr_row, column=i + 1, value=label)
        cell.font = _bold_font()
        cell.border = thin
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows — sorted by (issue_date, type, invoice_number)
    sorted_invoices = sorted(invoices, key=_invoice_sort_key)
    data_row = hdr_row + 1
    for idx, inv in enumerate(sorted_invoices, start=1):
        item_count = len(inv.items)
        type_label = TYPE_LABEL_EN.get(inv.type.value, inv.type.value.title())
        values = [
            idx,
            inv.issue_date,
            type_label,
            inv.recipient_name,
            item_count,
            float(inv.total_amount),
        ]
        for col_idx, val in enumerate(values):
            cell = ws.cell(row=data_row, column=col_idx + 1, value=val)
            cell.border = thin
            if col_idx == 5:  # Total column — currency
                cell.number_format = EUR_FR_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (0, 4):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 1:
                cell.number_format = "YYYY-MM-DD"
        data_row += 1

    # GRAND TOTAL band
    grand_row = data_row
    grand_total = sum((inv.total_amount for inv in invoices), Decimal("0"))
    for col in range(1, len(_INVOICE_HEADERS) + 1):
        cell = ws.cell(row=grand_row, column=col)
        cell.border = thick
        cell.fill = grand_fill
        cell.font = _bold_font()

    ws.cell(row=grand_row, column=1, value="GRAND TOTAL").font = _bold_font()
    ws.cell(row=grand_row, column=1).border = thick
    ws.cell(row=grand_row, column=1).fill = grand_fill

    total_cell = ws.cell(row=grand_row, column=len(_INVOICE_HEADERS), value=float(grand_total))
    total_cell.number_format = EUR_FR_FORMAT
    total_cell.font = _bold_font()
    total_cell.border = thick
    total_cell.fill = grand_fill
    total_cell.alignment = Alignment(horizontal="right")

    return grand_row + 1


# ---------------------------------------------------------------------------
# Per-type sheet writer
# ---------------------------------------------------------------------------


def _write_type_sheet(
    ws: Worksheet, context: InvoiceExportContext, invoice_type: InvoiceType, invoices: List[Invoice]
) -> None:
    """Write header band + invoice table + footer total for a single type sheet."""
    thin = _thin_border()
    thick = _thick_top_border()
    fill = _header_fill()
    type_label = TYPE_LABEL_EN.get(invoice_type.value, invoice_type.value.title())

    from_label = context.range.from_month.strftime("%Y-%m")
    to_label = context.range.to_month.strftime("%Y-%m")

    # Row 1 — title
    _merge_and_write(
        ws,
        row=1,
        start_col=1,
        end_col=5,
        value=f"{type_label} invoices — {context.project_name}",
        font=Font(bold=True, size=13),
    )

    # Row 2 — meta
    meta = f"Range: {from_label} to {to_label}"
    ws.cell(row=2, column=1, value=meta).font = _italic_font(9)

    # Row 3 blank → header at row 4
    hdr_row = 4
    for i, label in enumerate(_TYPE_SHEET_HEADERS):
        cell = ws.cell(row=hdr_row, column=i + 1, value=label)
        cell.font = _bold_font()
        cell.border = thin
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    sorted_invoices = sorted(invoices, key=_invoice_sort_key)
    data_row = hdr_row + 1
    for idx, inv in enumerate(sorted_invoices, start=1):
        item_count = len(inv.items)
        values = [idx, inv.issue_date, inv.recipient_name, item_count, float(inv.total_amount)]
        for col_idx, val in enumerate(values):
            cell = ws.cell(row=data_row, column=col_idx + 1, value=val)
            cell.border = thin
            if col_idx == 4:  # Total
                cell.number_format = EUR_FR_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (0, 3):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 1:
                cell.number_format = "YYYY-MM-DD"
        data_row += 1

    # Footer total row
    footer_row = data_row
    type_total = sum((inv.total_amount for inv in invoices), Decimal("0"))
    for col in range(1, len(_TYPE_SHEET_HEADERS) + 1):
        cell = ws.cell(row=footer_row, column=col)
        cell.border = thick
        cell.font = _bold_font()

    ws.cell(row=footer_row, column=1, value="TOTAL").font = _bold_font()
    ws.cell(row=footer_row, column=1).border = thick

    total_cell = ws.cell(row=footer_row, column=len(_TYPE_SHEET_HEADERS), value=float(type_total))
    total_cell.number_format = EUR_FR_FORMAT
    total_cell.font = _bold_font()
    total_cell.border = thick
    total_cell.alignment = Alignment(horizontal="right")

    _set_col_widths(ws, _TYPE_SHEET_COL_WIDTHS)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_xlsx(context: InvoiceExportContext, bundle: InvoiceBundle) -> bytes:
    """Generate an xlsx workbook for invoice export and return its raw bytes.

    Sheet 1 "Summary": header band, KPI table, subtotals by type, invoices table,
    grand total band.

    Sheets 2..k: one sheet per InvoiceType present in the bundle (skip empty types).
    Sheet titles: "Client invoices" / "Labor invoices" / "Supplier invoices".

    Empty range: only the Summary sheet, with an italic "No invoices in range" line
    replacing KPI/tables.

    Currency cells store raw float values; number_format renders as fr-FR EUR in Excel.
    """
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    next_row = _write_summary_header_band(ws_summary, context)

    if bundle.invoice_count == 0:
        from_label = context.range.from_month.strftime("%Y-%m")
        to_label = context.range.to_month.strftime("%Y-%m")
        empty_cell = ws_summary.cell(row=next_row, column=1, value=f"No invoices in range {from_label} to {to_label}")
        empty_cell.font = Font(italic=True)
        _set_col_widths(ws_summary, _INVOICE_COL_WIDTHS)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # KPI table
    next_row = _write_kpi_table(ws_summary, start_row=next_row, context=context, bundle=bundle)

    # Subtotals by type
    next_row = _write_subtotals_section(ws_summary, start_row=next_row, subtotals=bundle.subtotals_by_type)

    # Invoices table + grand total
    _write_invoices_section(ws_summary, start_row=next_row, invoices=bundle.invoices)

    _set_col_widths(ws_summary, _INVOICE_COL_WIDTHS)

    # Per-type sheets — one per type that has invoices (maintain canonical order)
    for invoice_type in (InvoiceType.CLIENT, InvoiceType.LABOR, InvoiceType.SUPPLIER):
        type_invoices = [inv for inv in bundle.invoices if inv.type == invoice_type]
        if not type_invoices:
            continue
        type_label = TYPE_LABEL_EN.get(invoice_type.value, invoice_type.value.title())
        sheet_title = f"{type_label} invoices"
        ws_type = wb.create_sheet(title=sheet_title)
        _write_type_sheet(ws_type, context, invoice_type, type_invoices)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
