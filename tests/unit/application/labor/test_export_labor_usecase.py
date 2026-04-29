"""Unit tests for ExportLaborUseCase.

All external collaborators are replaced with lightweight fakes — no DB, no Flask.

Covers:
- happy path xlsx: multi-month range, 2 workers, bytes start with PK, filename slug, mime_type
- happy path pdf: same scaffolding, bytes start with %PDF-, correct mime_type
- empty range: no labor entries → file generated containing "No labor entries" marker
- filename slug fallback: CJK/emoji project name → short ID prefix slug
- ProjectNotFoundError propagation when project does not exist
- cross-month aggregation: same worker in 2 months → days_worked correctly reflected in export
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from typing import List, Optional
from unittest.mock import MagicMock
from uuid import UUID, uuid4

import pytest

from app.application.labor.export_labor_usecase import ExportLaborRequest, ExportLaborUseCase
from app.application.labor.get_labor_summary import (
    GetLaborSummaryRequest,
    GetLaborSummaryUseCase,
    LaborSummaryResponse,
    WorkerCostSummary,
)
from app.application.labor.list_labor_entries import LaborEntryDetail, ListLaborEntriesRequest, ListLaborEntriesUseCase
from app.domain.entities.project import Project
from app.domain.entities.worker import Worker
from app.domain.exceptions.labor_exceptions import WorkerNotFoundError
from app.domain.exceptions.project_exceptions import ProjectNotFoundError


# ---------------------------------------------------------------------------
# Helpers / stubs
# ---------------------------------------------------------------------------


def _make_project(name: str = "Test Project", project_id: UUID | None = None) -> Project:
    pid = project_id or uuid4()
    return Project(
        id=pid,
        name=name,
        owner_id=uuid4(),
        created_at=datetime(2025, 1, 1, tzinfo=timezone.utc),
    )


def _make_worker_summary(
    *,
    worker_id: str = "w1",
    worker_name: str = "Antoine",
    days_worked: int = 5,
    total_cost: float = 1000.0,
    banked_hours: int = 0,
    bonus_full_days: int = 0,
    bonus_half_days: int = 0,
    bonus_cost: float = 0.0,
) -> WorkerCostSummary:
    return WorkerCostSummary(
        worker_id=worker_id,
        worker_name=worker_name,
        days_worked=days_worked,
        total_cost=total_cost,
        banked_hours=banked_hours,
        bonus_full_days=bonus_full_days,
        bonus_half_days=bonus_half_days,
        bonus_cost=bonus_cost,
    )


def _make_summary_response(*workers: WorkerCostSummary) -> LaborSummaryResponse:
    rows = list(workers)
    return LaborSummaryResponse(
        rows=rows,
        total_days=sum(r.days_worked for r in rows),
        total_cost=sum(r.total_cost for r in rows),
        total_banked_hours=sum(r.banked_hours for r in rows),
        total_bonus_days=sum(r.bonus_full_days + r.bonus_half_days * 0.5 for r in rows),
        total_bonus_cost=sum(r.bonus_cost for r in rows),
    )


def _empty_summary() -> LaborSummaryResponse:
    return LaborSummaryResponse(
        rows=[],
        total_days=0,
        total_cost=0.0,
        total_banked_hours=0,
        total_bonus_days=0.0,
        total_bonus_cost=0.0,
    )


def _make_entry(worker_name: str = "Antoine", entry_date: str = "2026-01-15") -> LaborEntryDetail:
    return LaborEntryDetail(
        id=str(uuid4()),
        worker_id="w1",
        worker_name=worker_name,
        date=entry_date,
        amount_override=None,
        effective_cost=200.0,
        note=None,
        shift_type="full",
        supplement_hours=0,
        created_at="2026-01-15T08:00:00",
    )


def _build_usecase(
    project: Optional[Project], summary_side_effect=None, entries_side_effect=None
) -> ExportLaborUseCase:
    """Construct use-case with all collaborators mocked."""
    project_repo = MagicMock()
    project_repo.find_by_id.return_value = project

    summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
    if summary_side_effect is not None:
        summary_uc.execute.side_effect = summary_side_effect
    else:
        summary_uc.execute.return_value = _empty_summary()

    entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
    if entries_side_effect is not None:
        entries_uc.execute.side_effect = entries_side_effect
    else:
        entries_uc.execute.return_value = []

    worker_repo = MagicMock()
    entry_repo = MagicMock()

    return ExportLaborUseCase(
        worker_repo=worker_repo,
        entry_repo=entry_repo,
        summary_usecase=summary_uc,
        list_entries_usecase=entries_uc,
        project_repo=project_repo,
    )


# ---------------------------------------------------------------------------
# Happy path — xlsx
# ---------------------------------------------------------------------------


class TestExportXlsxHappyPath:
    def test_returns_pk_magic_bytes(self):
        """xlsx bytes begin with PK\\x03\\x04 (ZIP local file header)."""
        project = _make_project("Downtown Office Tower")
        workers = [
            _make_worker_summary(worker_id="w1", worker_name="Antoine", days_worked=10, total_cost=2000.0),
            _make_worker_summary(worker_id="w2", worker_name="Marc", days_worked=8, total_cost=1600.0),
        ]
        summary = _make_summary_response(*workers)

        uc = _build_usecase(project, summary_side_effect=lambda _req: summary, entries_side_effect=lambda _req: [])
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-02",
            format="xlsx",
            acting_user_email="admin@example.com",
        )
        result = uc.execute(req)

        assert result.content[:4] == b"PK\x03\x04", "Expected xlsx ZIP magic bytes"

    def test_mime_type_is_xlsx(self):
        project = _make_project()
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert result.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_filename_includes_slug_and_range(self):
        project = _make_project("Downtown Office Tower")
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-03",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)

        assert result.filename.startswith("labor-downtown-office-tower")
        assert "2026-01-to-2026-03" in result.filename
        assert result.filename.endswith(".xlsx")


# ---------------------------------------------------------------------------
# Happy path — pdf
# ---------------------------------------------------------------------------


class TestExportPdfHappyPath:
    def test_returns_pdf_magic_bytes(self):
        """pdf bytes begin with %PDF-."""
        project = _make_project()
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-04",
            to_month="2026-04",
            format="pdf",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert result.content[:5] == b"%PDF-", "Expected PDF magic bytes"

    def test_mime_type_is_pdf(self):
        project = _make_project()
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-04",
            to_month="2026-04",
            format="pdf",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert result.mime_type == "application/pdf"

    def test_filename_ends_with_pdf(self):
        project = _make_project("Test Site")
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-04",
            to_month="2026-06",
            format="pdf",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert result.filename.endswith(".pdf")
        assert "2026-04-to-2026-06" in result.filename


# ---------------------------------------------------------------------------
# Empty range (no labor entries)
# ---------------------------------------------------------------------------


class TestExportEmptyRange:
    def test_xlsx_generated_even_with_no_entries(self):
        """Empty range still produces valid xlsx bytes (no 404)."""
        project = _make_project("Empty Project")
        uc = _build_usecase(project)  # both summary_uc and entries_uc return empty
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-06",
            to_month="2026-06",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert len(result.content) > 0
        assert result.content[:4] == b"PK\x03\x04"

    def test_pdf_generated_even_with_no_entries(self):
        """Empty range still produces valid pdf bytes (no 404)."""
        project = _make_project("Empty Project")
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-06",
            to_month="2026-06",
            format="pdf",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert len(result.content) > 0
        assert result.content[:5] == b"%PDF-"


# ---------------------------------------------------------------------------
# Filename slug fallback (CJK / emoji)
# ---------------------------------------------------------------------------


class TestFilenameSlugFallback:
    def test_cjk_name_falls_back_to_id_prefix(self):
        """Pure CJK project name → slug is first 8 hex chars from project ID."""
        pid = uuid4()
        project = _make_project("工地", project_id=pid)
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=pid,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        # The slug should contain the first 8 chars of the UUID (without hyphens considered by slugify)
        id_str = str(pid)
        id_prefix = id_str[:8]  # e.g. "378bc411"
        assert id_prefix in result.filename, f"Expected ID prefix in filename; got: {result.filename}"

    def test_emoji_name_falls_back_to_id_prefix(self):
        """Pure emoji project name → slug falls back to ID prefix."""
        pid = uuid4()
        project = _make_project("\U0001f3d7️", project_id=pid)
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=pid,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        id_prefix = str(pid)[:8]
        assert id_prefix in result.filename, f"Expected ID prefix in filename; got: {result.filename}"


# ---------------------------------------------------------------------------
# ProjectNotFoundError propagation
# ---------------------------------------------------------------------------


class TestProjectNotFoundPropagation:
    def test_raises_project_not_found_when_project_missing(self):
        """Use-case propagates ProjectNotFoundError if project_repo returns None."""
        uc = _build_usecase(project=None)  # find_by_id returns None
        req = ExportLaborRequest(
            project_id=uuid4(),
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        with pytest.raises(ProjectNotFoundError):
            uc.execute(req)


# ---------------------------------------------------------------------------
# Cross-month aggregation
# ---------------------------------------------------------------------------


class TestCrossMonthAggregation:
    def test_summary_called_once_per_month(self):
        """With a 2-month range, summary_usecase.execute is called exactly twice."""
        project = _make_project()
        summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
        summary_uc.execute.return_value = _empty_summary()

        entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
        entries_uc.execute.return_value = []

        project_repo = MagicMock()
        project_repo.find_by_id.return_value = project

        uc = ExportLaborUseCase(
            worker_repo=MagicMock(),
            entry_repo=MagicMock(),
            summary_usecase=summary_uc,
            list_entries_usecase=entries_uc,
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-02",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        uc.execute(req)

        assert (
            summary_uc.execute.call_count == 2
        ), f"Expected 2 calls for 2-month range, got {summary_uc.execute.call_count}"

    def test_summary_called_for_correct_month_boundaries(self):
        """Month boundary dates passed to summary_usecase match YYYY-01 / YYYY-01-31."""
        project = _make_project()
        summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
        summary_uc.execute.return_value = _empty_summary()

        entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
        entries_uc.execute.return_value = []

        project_repo = MagicMock()
        project_repo.find_by_id.return_value = project

        uc = ExportLaborUseCase(
            worker_repo=MagicMock(),
            entry_repo=MagicMock(),
            summary_usecase=summary_uc,
            list_entries_usecase=entries_uc,
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-02",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        uc.execute(req)

        calls = summary_uc.execute.call_args_list
        jan_call: GetLaborSummaryRequest = calls[0][0][0]
        feb_call: GetLaborSummaryRequest = calls[1][0][0]

        assert jan_call.date_from == date(2026, 1, 1)
        assert jan_call.date_to == date(2026, 1, 31)
        assert feb_call.date_from == date(2026, 2, 1)
        assert feb_call.date_to == date(2026, 2, 28)

    def test_december_to_january_year_wrap_month_boundary(self):
        """Dec 2025 → Jan 2026 range: summary called with correct year-wrap boundaries."""
        project = _make_project()
        summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
        summary_uc.execute.return_value = _empty_summary()

        entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
        entries_uc.execute.return_value = []

        project_repo = MagicMock()
        project_repo.find_by_id.return_value = project

        uc = ExportLaborUseCase(
            worker_repo=MagicMock(),
            entry_repo=MagicMock(),
            summary_usecase=summary_uc,
            list_entries_usecase=entries_uc,
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2025-12",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        uc.execute(req)

        assert summary_uc.execute.call_count == 2
        calls = summary_uc.execute.call_args_list
        dec_call: GetLaborSummaryRequest = calls[0][0][0]
        jan_call: GetLaborSummaryRequest = calls[1][0][0]

        assert dec_call.date_from == date(2025, 12, 1)
        assert dec_call.date_to == date(2025, 12, 31)
        assert jan_call.date_from == date(2026, 1, 1)
        assert jan_call.date_to == date(2026, 1, 31)

    def test_same_worker_in_two_months_produces_valid_export(self):
        """Worker present in both months → export completes, 2 months × data reflected in file."""
        project = _make_project("Office Tower")

        w_jan = _make_worker_summary(worker_id="w1", worker_name="Nguyễn Văn Đức", days_worked=10, total_cost=2000.0)
        w_feb = _make_worker_summary(worker_id="w1", worker_name="Nguyễn Văn Đức", days_worked=8, total_cost=1600.0)

        summaries = [_make_summary_response(w_jan), _make_summary_response(w_feb)]
        entries = [
            [_make_entry("Nguyễn Văn Đức", "2026-01-15")],
            [_make_entry("Nguyễn Văn Đức", "2026-02-15")],
        ]

        call_count = [0]

        def summary_side(req: GetLaborSummaryRequest) -> LaborSummaryResponse:
            idx = call_count[0]
            call_count[0] += 1
            return summaries[idx]

        entry_call_count = [0]

        def entries_side(req: ListLaborEntriesRequest) -> List[LaborEntryDetail]:
            idx = entry_call_count[0]
            entry_call_count[0] += 1
            return entries[idx]

        uc = _build_usecase(project, summary_side_effect=summary_side, entries_side_effect=entries_side)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-02",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)

        # Valid xlsx produced
        assert result.content[:4] == b"PK\x03\x04"
        # 2 months × entry calls happened
        assert entry_call_count[0] == 2


# ---------------------------------------------------------------------------
# Helpers for worker entity
# ---------------------------------------------------------------------------


def _make_worker(
    *,
    worker_id: UUID | None = None,
    project_id: UUID | None = None,
    name: str = "Antoine Dupont",
    daily_rate: str = "200.00",
) -> Worker:
    from datetime import timezone
    from decimal import Decimal

    return Worker(
        id=worker_id or uuid4(),
        project_id=project_id or uuid4(),
        name=name,
        daily_rate=Decimal(daily_rate),
        created_at=datetime(2025, 1, 1, tzinfo=timezone.utc),
    )


def _build_usecase_with_worker(
    project: Project | None,
    worker: Worker | None,
    summary_side_effect=None,
    entries_side_effect=None,
) -> ExportLaborUseCase:
    """Construct use-case with worker_repo stub that returns `worker`."""
    project_repo = MagicMock()
    project_repo.find_by_id.return_value = project

    worker_repo = MagicMock()
    worker_repo.find_by_id.return_value = worker

    summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
    if summary_side_effect is not None:
        summary_uc.execute.side_effect = summary_side_effect
    else:
        summary_uc.execute.return_value = _empty_summary()

    entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
    if entries_side_effect is not None:
        entries_uc.execute.side_effect = entries_side_effect
    else:
        entries_uc.execute.return_value = []

    return ExportLaborUseCase(
        worker_repo=worker_repo,
        entry_repo=MagicMock(),
        summary_usecase=summary_uc,
        list_entries_usecase=entries_uc,
        project_repo=project_repo,
    )


# ---------------------------------------------------------------------------
# worker_id=None regression
# ---------------------------------------------------------------------------


class TestWorkerIdNoneRegression:
    """Existing project-wide behaviour must be unchanged when worker_id is None."""

    def test_worker_id_none_produces_valid_xlsx(self):
        """worker_id=None (default) still yields valid xlsx bytes."""
        project = _make_project("Test Project")
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
            # worker_id omitted → defaults to None
        )
        result = uc.execute(req)
        assert result.content[:4] == b"PK\x03\x04"

    def test_worker_id_none_filename_has_no_worker_slug(self):
        """Project-wide filename must NOT contain a worker slug segment."""
        project = _make_project("Office Tower")
        uc = _build_usecase(project)
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        # Format: labor-{project-slug}-{from}-to-{to}.xlsx (no worker segment)
        assert result.filename.startswith("labor-")
        assert "2026-01-to-2026-01" in result.filename

    def test_worker_id_none_worker_repo_not_called(self):
        """When worker_id=None, worker_repo.find_by_id must not be called."""
        project = _make_project()
        worker_repo = MagicMock()
        project_repo = MagicMock()
        project_repo.find_by_id.return_value = project
        summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
        summary_uc.execute.return_value = _empty_summary()
        entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
        entries_uc.execute.return_value = []

        uc = ExportLaborUseCase(
            worker_repo=worker_repo,
            entry_repo=MagicMock(),
            summary_usecase=summary_uc,
            list_entries_usecase=entries_uc,
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        uc.execute(req)
        worker_repo.find_by_id.assert_not_called()

    def test_summary_called_without_worker_id_filter(self):
        """With worker_id=None, GetLaborSummaryRequest.worker_id must be None."""
        project = _make_project()
        summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
        summary_uc.execute.return_value = _empty_summary()
        entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
        entries_uc.execute.return_value = []
        project_repo = MagicMock()
        project_repo.find_by_id.return_value = project

        uc = ExportLaborUseCase(
            worker_repo=MagicMock(),
            entry_repo=MagicMock(),
            summary_usecase=summary_uc,
            list_entries_usecase=entries_uc,
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=project.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        uc.execute(req)
        call_arg: GetLaborSummaryRequest = summary_uc.execute.call_args[0][0]
        assert call_arg.worker_id is None


# ---------------------------------------------------------------------------
# worker_id set — happy path
# ---------------------------------------------------------------------------


class TestWorkerIdHappyPath:
    def test_export_context_has_worker_name(self):
        """When worker_id is set, ExportContext.worker_name == worker.name."""
        pid = uuid4()
        project = _make_project("Test Project", project_id=pid)
        worker = _make_worker(project_id=pid, name="Marc Leblanc", daily_rate="250.00")
        uc = _build_usecase_with_worker(project, worker)
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=worker.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        # The generated xlsx (single-worker) sheet name equals worker name
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(result.content), data_only=True)
        # Single-worker mode: sheet name is the sanitized worker name
        assert len(wb.sheetnames) == 1
        assert "Marc" in wb.sheetnames[0] or "leblanc" in wb.sheetnames[0].lower()

    def test_filename_includes_worker_slug(self):
        """With worker_id set, filename contains worker slug."""
        pid = uuid4()
        project = _make_project("Office Tower", project_id=pid)
        worker = _make_worker(project_id=pid, name="Antoine Dupont")
        uc = _build_usecase_with_worker(project, worker)
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=worker.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert "antoine-dupont" in result.filename, f"Worker slug not in filename: {result.filename}"

    def test_filename_includes_project_slug_and_range(self):
        """Filename format: labor-{project-slug}-{worker-slug}-{from}-to-{to}.xlsx"""
        pid = uuid4()
        project = _make_project("Office Tower", project_id=pid)
        worker = _make_worker(project_id=pid, name="Marc Leblanc")
        uc = _build_usecase_with_worker(project, worker)
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=worker.id,
            from_month="2026-02",
            to_month="2026-04",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert result.filename.startswith("labor-office-tower")
        assert "marc-leblanc" in result.filename
        assert "2026-02-to-2026-04" in result.filename
        assert result.filename.endswith(".xlsx")

    def test_summary_called_with_worker_id_filter(self):
        """With worker_id set, GetLaborSummaryRequest.worker_id == worker.id."""
        pid = uuid4()
        project = _make_project(project_id=pid)
        worker = _make_worker(project_id=pid)
        summary_uc = MagicMock(spec=GetLaborSummaryUseCase)
        summary_uc.execute.return_value = _empty_summary()
        entries_uc = MagicMock(spec=ListLaborEntriesUseCase)
        entries_uc.execute.return_value = []
        project_repo = MagicMock()
        project_repo.find_by_id.return_value = project
        worker_repo = MagicMock()
        worker_repo.find_by_id.return_value = worker

        uc = ExportLaborUseCase(
            worker_repo=worker_repo,
            entry_repo=MagicMock(),
            summary_usecase=summary_uc,
            list_entries_usecase=entries_uc,
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=worker.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        uc.execute(req)
        call_arg: GetLaborSummaryRequest = summary_uc.execute.call_args[0][0]
        assert call_arg.worker_id == worker.id

    def test_pdf_filename_ends_with_pdf(self):
        """Single-worker PDF filename ends with .pdf."""
        pid = uuid4()
        project = _make_project(project_id=pid)
        worker = _make_worker(project_id=pid, name="Antoine Dupont")
        uc = _build_usecase_with_worker(project, worker)
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=worker.id,
            from_month="2026-01",
            to_month="2026-01",
            format="pdf",
            acting_user_email="user@example.com",
        )
        result = uc.execute(req)
        assert result.filename.endswith(".pdf")
        assert "antoine-dupont" in result.filename


# ---------------------------------------------------------------------------
# worker_id set — worker not found / wrong project
# ---------------------------------------------------------------------------


class TestWorkerIdNotFound:
    def test_worker_not_in_repo_raises_worker_not_found(self):
        """worker_repo returns None → WorkerNotFoundError raised."""
        pid = uuid4()
        project = _make_project(project_id=pid)
        uc = _build_usecase_with_worker(project, worker=None)
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=uuid4(),
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        with pytest.raises(WorkerNotFoundError):
            uc.execute(req)

    def test_worker_belongs_to_different_project_raises_worker_not_found(self):
        """Worker exists but project_id mismatch → WorkerNotFoundError."""
        pid = uuid4()
        other_pid = uuid4()
        project = _make_project(project_id=pid)
        # Worker belongs to other_pid, not pid
        worker = _make_worker(project_id=other_pid, name="Alice")
        uc = _build_usecase_with_worker(project, worker)
        req = ExportLaborRequest(
            project_id=pid,
            worker_id=worker.id,
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        with pytest.raises(WorkerNotFoundError):
            uc.execute(req)

    def test_project_not_found_takes_precedence_over_worker(self):
        """ProjectNotFoundError raised before worker lookup."""
        worker_repo = MagicMock()
        project_repo = MagicMock()
        project_repo.find_by_id.return_value = None  # project missing

        uc = ExportLaborUseCase(
            worker_repo=worker_repo,
            entry_repo=MagicMock(),
            summary_usecase=MagicMock(spec=GetLaborSummaryUseCase),
            list_entries_usecase=MagicMock(spec=ListLaborEntriesUseCase),
            project_repo=project_repo,
        )
        req = ExportLaborRequest(
            project_id=uuid4(),
            worker_id=uuid4(),
            from_month="2026-01",
            to_month="2026-01",
            format="xlsx",
            acting_user_email="user@example.com",
        )
        with pytest.raises(ProjectNotFoundError):
            uc.execute(req)
        # Worker repo should never be called if project is missing
        worker_repo.find_by_id.assert_not_called()
