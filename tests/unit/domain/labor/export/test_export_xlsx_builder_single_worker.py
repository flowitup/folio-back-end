"""Unit tests for xlsx_builder._build_xlsx_single_worker (single-worker mode).

Covers:
- One sheet, sheet name = sanitized worker name (≤31 chars, no forbidden chars)
- Header rows include 'Worker:' and 'Rate:' text
- Monthly summary table present (one row per month bucket with entries)
- Daily detail table present, sorted ascending by date
- Empty range → 'No labor entries in range …' cell only
- _sanitize_sheet_name edge cases (forbidden chars stripped, length capped)
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import openpyxl

from app.application.labor.get_labor_summary import LaborSummaryResponse, WorkerCostSummary
from app.application.labor.list_labor_entries import LaborEntryDetail
from app.domain.labor.export.models import ExportContext, ExportRange, MonthBucket
from app.domain.labor.export.xlsx_builder import _sanitize_sheet_name, build_xlsx


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load(raw: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(BytesIO(raw), data_only=True)


def _make_worker_context(
    worker_name: str = "Antoine Dupont",
    worker_daily_rate: Decimal | None = Decimal("200.00"),
    from_month: date = date(2026, 4, 1),
    to_month: date = date(2026, 4, 30),
) -> ExportContext:
    return ExportContext(
        project_name="Downtown Office Tower",
        project_id=uuid4(),
        range=ExportRange(from_month=from_month, to_month=to_month),
        generated_at=datetime(2026, 4, 28, 17, 0, 0),
        generated_by_email="admin@example.com",
        worker_name=worker_name,
        worker_daily_rate=worker_daily_rate,
    )


def _make_worker_summary(
    *,
    worker_id: str = "w1",
    worker_name: str = "Antoine Dupont",
    days_worked: int = 10,
    total_cost: float = 2000.0,
    banked_hours: int = 0,
    bonus_full_days: int = 0,
    bonus_half_days: int = 0,
    bonus_cost: float = 0.0,
) -> WorkerCostSummary:
    return WorkerCostSummary(
        worker_id=worker_id,
        worker_name=worker_name,
        days_worked=days_worked,
        total_cost=total_cost,
        banked_hours=banked_hours,
        bonus_full_days=bonus_full_days,
        bonus_half_days=bonus_half_days,
        bonus_cost=bonus_cost,
    )


def _make_summary_response(*workers: WorkerCostSummary) -> LaborSummaryResponse:
    rows = list(workers)
    return LaborSummaryResponse(
        rows=rows,
        total_days=sum(r.days_worked for r in rows),
        total_cost=sum(r.total_cost for r in rows),
        total_banked_hours=sum(r.banked_hours for r in rows),
        total_bonus_days=sum(r.bonus_full_days + r.bonus_half_days * 0.5 for r in rows),
        total_bonus_cost=sum(r.bonus_cost for r in rows),
    )


def _make_entry(
    *,
    worker_id: str = "w1",
    worker_name: str = "Antoine Dupont",
    entry_date: str = "2026-04-15",
    shift_type: str = "full",
    effective_cost: float = 200.0,
    supplement_hours: int = 0,
    amount_override: float | None = None,
    note: str | None = None,
) -> LaborEntryDetail:
    return LaborEntryDetail(
        id=str(uuid4()),
        worker_id=worker_id,
        worker_name=worker_name,
        date=entry_date,
        amount_override=amount_override,
        effective_cost=effective_cost,
        note=note,
        shift_type=shift_type,
        supplement_hours=supplement_hours,
        created_at="2026-04-28T17:00:00",
    )


def _make_single_worker_bucket(
    month: date = date(2026, 4, 1),
    *,
    worker_name: str = "Antoine Dupont",
    days_worked: int = 10,
    entries: list[LaborEntryDetail] | None = None,
) -> MonthBucket:
    w = _make_worker_summary(worker_name=worker_name, days_worked=days_worked)
    return MonthBucket(
        month=month,
        summary=_make_summary_response(w),
        daily_entries=entries or [_make_entry(worker_name=worker_name)],
    )


# ---------------------------------------------------------------------------
# Sheet structure
# ---------------------------------------------------------------------------


class TestSingleWorkerSheetCount:
    def test_one_sheet_only(self):
        """Single-worker mode produces exactly one sheet."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket()
        wb = _load(build_xlsx(ctx, [bucket]))
        assert len(wb.sheetnames) == 1, f"Expected 1 sheet, got {wb.sheetnames}"

    def test_sheet_name_is_sanitized_worker_name(self):
        """Sheet name equals sanitized worker name (≤31 chars)."""
        ctx = _make_worker_context(worker_name="Antoine Dupont")
        bucket = _make_single_worker_bucket()
        wb = _load(build_xlsx(ctx, [bucket]))
        assert wb.sheetnames[0] == "Antoine Dupont"

    def test_sheet_name_length_capped_at_31(self):
        """Worker name longer than 31 chars → sheet name capped at 31."""
        long_name = "Jean-Baptiste Emmanuel Zorg de la Bourboule"
        ctx = _make_worker_context(worker_name=long_name)
        bucket = _make_single_worker_bucket(worker_name=long_name)
        wb = _load(build_xlsx(ctx, [bucket]))
        assert len(wb.sheetnames[0]) <= 31

    def test_sheet_name_forbidden_chars_stripped(self):
        """Forbidden Excel chars []:*?/\\ stripped from sheet name."""
        ctx = _make_worker_context(worker_name="Worker [Main] / Backup")
        bucket = _make_single_worker_bucket(worker_name="Worker [Main] / Backup")
        wb = _load(build_xlsx(ctx, [bucket]))
        sheet_name = wb.sheetnames[0]
        for ch in ["[", "]", ":", "*", "?", "/", "\\"]:
            assert ch not in sheet_name, f"Forbidden char '{ch}' found in sheet name: {sheet_name!r}"


# ---------------------------------------------------------------------------
# Header block
# ---------------------------------------------------------------------------


class TestSingleWorkerHeaderBlock:
    def test_header_row1_title(self):
        """A1 contains 'Folio · Labor Export'."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        assert ws["A1"].value == "Folio · Labor Export"

    def test_header_row2_project_name(self):
        """A2 contains the project name."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        assert "Downtown Office Tower" in (ws["A2"].value or "")

    def test_header_row4_generated_by_email(self):
        """A4 contains generated_by_email."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        assert "admin@example.com" in (ws["A4"].value or "")

    def test_header_contains_worker_line(self):
        """A cell in header area contains 'Worker: Antoine Dupont'."""
        ctx = _make_worker_context(worker_name="Antoine Dupont")
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        # Row 5 is the worker sub-header (after blank row 5 from _write_header_block)
        worker_line = ws.cell(row=5, column=1).value or ""
        assert "Worker:" in worker_line, f"'Worker:' not found in row 5: {worker_line!r}"
        assert "Antoine Dupont" in worker_line

    def test_header_contains_rate_line(self):
        """Worker sub-header row also contains 'Rate:' with the daily rate."""
        ctx = _make_worker_context(worker_daily_rate=Decimal("250.00"))
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        worker_line = ws.cell(row=5, column=1).value or ""
        assert "Rate:" in worker_line, f"'Rate:' not found in row 5: {worker_line!r}"
        assert "250" in worker_line

    def test_header_rate_em_dash_when_none(self):
        """When worker_daily_rate is None, rate shows '—'."""
        ctx = _make_worker_context(worker_daily_rate=None)
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        worker_line = ws.cell(row=5, column=1).value or ""
        assert "—" in worker_line


# ---------------------------------------------------------------------------
# Monthly summary table
# ---------------------------------------------------------------------------


class TestSingleWorkerMonthlySummary:
    def test_monthly_summary_table_present(self):
        """A summary table with 'Worker' column header exists."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        all_values = [ws.cell(row=r, column=1).value for r in range(1, 40)]
        assert "Worker" in all_values, f"'Worker' header not found in col A. Values: {all_values}"

    def test_monthly_summary_row_shows_worker_name(self):
        """Data row in summary table contains worker name."""
        ctx = _make_worker_context(worker_name="Marc Leblanc")
        bucket = _make_single_worker_bucket(worker_name="Marc Leblanc")
        ws = _load(build_xlsx(ctx, [bucket])).active
        all_values = [ws.cell(row=r, column=1).value for r in range(1, 50)]
        assert "Marc Leblanc" in all_values, f"Worker name not found. Col A values: {all_values}"

    def test_monthly_summary_days_correct(self):
        """Days column in summary row matches seeded days_worked."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket(days_worked=7)
        ws = _load(build_xlsx(ctx, [bucket])).active
        # Find 'Worker' header row; next row should be data row
        worker_hdr_row = None
        for r in range(1, 50):
            if ws.cell(row=r, column=1).value == "Worker":
                worker_hdr_row = r
                break
        assert worker_hdr_row is not None
        days_val = ws.cell(row=worker_hdr_row + 1, column=2).value
        assert days_val == 7, f"Expected 7 days, got {days_val}"


# ---------------------------------------------------------------------------
# Daily detail table
# ---------------------------------------------------------------------------


class TestSingleWorkerDailyDetail:
    def test_daily_detail_section_present(self):
        """'Daily detail' section label appears."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket()
        ws = _load(build_xlsx(ctx, [bucket])).active
        all_values = [ws.cell(row=r, column=1).value for r in range(1, 60)]
        assert "Daily detail" in all_values, f"'Daily detail' section not found. Values: {all_values}"

    def test_daily_detail_headers(self):
        """Daily detail table has Date, Worker, Shift, Effective cost headers."""
        ctx = _make_worker_context()
        bucket = _make_single_worker_bucket(entries=[_make_entry(entry_date="2026-04-10")])
        ws = _load(build_xlsx(ctx, [bucket])).active
        detail_row = None
        for r in range(1, 60):
            if ws.cell(row=r, column=1).value == "Daily detail":
                detail_row = r
                break
        assert detail_row is not None
        hdr_row = detail_row + 1
        headers = [ws.cell(row=hdr_row, column=c).value for c in range(1, 8)]
        assert headers[0] == "Date"
        assert headers[1] == "Worker"
        assert headers[2] == "Shift"
        assert headers[5] == "Effective cost"

    def test_daily_detail_sorted_ascending_by_date(self):
        """Daily detail entries are sorted ascending by date."""
        ctx = _make_worker_context()
        entries = [
            _make_entry(entry_date="2026-04-20", effective_cost=200.0),
            _make_entry(entry_date="2026-04-05", effective_cost=150.0),
            _make_entry(entry_date="2026-04-12", effective_cost=180.0),
        ]
        w = _make_worker_summary(days_worked=3)
        bucket = MonthBucket(
            month=date(2026, 4, 1),
            summary=_make_summary_response(w),
            daily_entries=entries,
        )
        ws = _load(build_xlsx(ctx, [bucket])).active
        detail_row = None
        for r in range(1, 70):
            if ws.cell(row=r, column=1).value == "Daily detail":
                detail_row = r
                break
        assert detail_row is not None
        data_start = detail_row + 2  # skip table header
        dates = [ws.cell(row=data_start + i, column=1).value for i in range(3)]
        assert dates == sorted(dates), f"Entries not sorted: {dates}"
        assert dates[0] == "2026-04-05"
        assert dates[2] == "2026-04-20"


# ---------------------------------------------------------------------------
# Multi-month layout
# ---------------------------------------------------------------------------


class TestSingleWorkerMultiMonth:
    def test_two_month_buckets_both_represented(self):
        """Two month buckets → both month section labels appear."""
        ctx = _make_worker_context(from_month=date(2026, 4, 1), to_month=date(2026, 5, 31))
        bucket_apr = _make_single_worker_bucket(month=date(2026, 4, 1), entries=[_make_entry(entry_date="2026-04-15")])
        bucket_may = _make_single_worker_bucket(month=date(2026, 5, 1), entries=[_make_entry(entry_date="2026-05-10")])
        ws = _load(build_xlsx(ctx, [bucket_may, bucket_apr])).active  # pass out of order
        all_values = [ws.cell(row=r, column=1).value for r in range(1, 100)]
        assert "Apr 2026" in all_values, f"'Apr 2026' not found. Values: {all_values}"
        assert "May 2026" in all_values, f"'May 2026' not found. Values: {all_values}"


# ---------------------------------------------------------------------------
# Empty range
# ---------------------------------------------------------------------------


class TestSingleWorkerEmptyRange:
    def test_empty_buckets_list_shows_message(self):
        """Zero buckets → 'No labor entries in range' message on single sheet."""
        ctx = _make_worker_context()
        wb = _load(build_xlsx(ctx, []))
        assert len(wb.sheetnames) == 1
        ws = wb.active
        all_values = [ws.cell(row=r, column=1).value for r in range(1, 20)]
        no_entries = any(v and "No labor entries in range" in str(v) for v in all_values)
        assert no_entries, f"Empty-state message not found. Values: {all_values}"

    def test_empty_bucket_with_no_rows_shows_message(self):
        """Bucket present but all empty → shows 'No labor entries in range'."""
        ctx = _make_worker_context()
        empty_bucket = MonthBucket(
            month=date(2026, 4, 1),
            summary=_make_summary_response(),  # no workers
            daily_entries=[],
        )
        wb = _load(build_xlsx(ctx, [empty_bucket]))
        ws = wb.active
        all_values = [ws.cell(row=r, column=1).value for r in range(1, 20)]
        no_entries = any(v and "No labor entries in range" in str(v) for v in all_values)
        assert no_entries, f"Empty-state message not found. Values: {all_values}"

    def test_empty_range_no_worker_table(self):
        """Empty range → no 'Worker' header in summary table (no data table rendered)."""
        ctx = _make_worker_context()
        wb = _load(build_xlsx(ctx, []))
        ws = wb.active
        # 'Worker:' sub-header (row 5) is fine; 'Worker' as table column header should NOT appear
        all_values = {ws.cell(row=r, column=1).value for r in range(6, 30)}
        # The table header "Worker" should not be present (only the sub-header at row 5)
        assert "Worker" not in all_values, f"'Worker' table header found in empty range output. Values: {all_values}"

    def test_empty_range_xlsx_magic_bytes(self):
        """Empty range still returns valid xlsx bytes."""
        ctx = _make_worker_context()
        raw = build_xlsx(ctx, [])
        assert raw[:4] == b"PK\x03\x04"


# ---------------------------------------------------------------------------
# _sanitize_sheet_name
# ---------------------------------------------------------------------------


class TestSanitizeSheetName:
    def test_normal_name_unchanged(self):
        result = _sanitize_sheet_name("Antoine Dupont", "fallback")
        assert result == "Antoine Dupont"

    def test_forbidden_chars_stripped(self):
        result = _sanitize_sheet_name("Worker[Main]:test?/\\*", "fallback")
        for ch in ["[", "]", ":", "*", "?", "/", "\\"]:
            assert ch not in result

    def test_empty_after_stripping_falls_back(self):
        """All-forbidden-char name → falls back to first 8 chars of fallback."""
        result = _sanitize_sheet_name("[]:*?/\\", "myworker")
        assert result == "myworker"

    def test_length_capped_at_31(self):
        long = "A" * 50
        result = _sanitize_sheet_name(long, "fallback")
        assert len(result) == 31

    def test_exactly_31_chars_preserved(self):
        name = "A" * 31
        result = _sanitize_sheet_name(name, "fallback")
        assert result == name

    def test_32_chars_truncated_to_31(self):
        name = "A" * 32
        result = _sanitize_sheet_name(name, "fallback")
        assert len(result) == 31
