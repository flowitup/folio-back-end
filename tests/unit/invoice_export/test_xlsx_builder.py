"""Unit tests for app.domain.invoice.export.xlsx_builder.build_xlsx.

Opens workbook via openpyxl.load_workbook(BytesIO(bytes)) to assert structure.

Covers:
- Summary sheet always present
- One sheet per type that has invoices (empty types skipped)
- Currency cells store float values with EUR_FR_FORMAT number_format
- Grand total row present in Summary invoices section
- Empty range produces only Summary sheet
- XSS-safe recipient name — openpyxl handles XML escaping without crash
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import openpyxl

from app.domain.entities.invoice import Invoice, InvoiceType
from app.domain.invoice.export.models import (
    InvoiceBundle,
    InvoiceExportContext,
    InvoiceExportRange,
    TypeSubtotal,
)
from app.domain.invoice.export.xlsx_builder import EUR_FR_FORMAT, build_xlsx
from app.domain.value_objects.invoice_item import InvoiceItem


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_context(project_name: str = "Test Project") -> InvoiceExportContext:
    return InvoiceExportContext(
        project_name=project_name,
        project_id=uuid4(),
        range=InvoiceExportRange(
            from_month=date(2026, 1, 1),
            to_month=date(2026, 3, 1),
        ),
        generated_at=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
        generated_by_email="admin@example.com",
        type_filter=None,
    )


def _make_invoice(
    *,
    project_id=None,
    invoice_type: InvoiceType = InvoiceType.RELEASED_FUNDS,
    amount: Decimal = Decimal("150.00"),
    recipient: str = "ACME Corp",
    issue_date: date = date(2026, 1, 15),
    invoice_number: str = "INV-001",
) -> Invoice:
    pid = project_id or uuid4()
    item = InvoiceItem(description="Service", quantity=Decimal("1"), unit_price=amount)
    return Invoice(
        id=uuid4(),
        project_id=pid,
        invoice_number=invoice_number,
        type=invoice_type,
        issue_date=issue_date,
        recipient_name=recipient,
        created_by=uuid4(),
        created_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
        updated_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
        items=[item],
    )


def _make_bundle(invoices: list, subtotals: list | None = None, grand_total: Decimal | None = None) -> InvoiceBundle:
    if subtotals is None:
        # Compute subtotals from invoices
        from app.domain.entities.invoice import InvoiceType as IT

        subtotals = []
        for t in (IT.RELEASED_FUNDS, IT.LABOR, IT.MATERIALS_SERVICES):
            scoped = [i for i in invoices if i.type == t]
            if scoped:
                subtotals.append(
                    TypeSubtotal(
                        type=t,
                        invoice_count=len(scoped),
                        total_amount=sum((i.total_amount for i in scoped), Decimal("0")),
                    )
                )
    gt = grand_total if grand_total is not None else sum((i.total_amount for i in invoices), Decimal("0"))
    return InvoiceBundle(
        invoices=invoices,
        subtotals_by_type=subtotals,
        grand_total=gt,
        invoice_count=len(invoices),
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_summary_sheet_present():
    """build_xlsx always produces a 'Summary' sheet as the first sheet."""
    ctx = _make_context()
    bundle = _make_bundle([_make_invoice()])
    wb = openpyxl.load_workbook(BytesIO(build_xlsx(ctx, bundle)))
    assert "Summary" in wb.sheetnames
    assert wb.sheetnames[0] == "Summary"


def test_one_sheet_per_type_skips_empty_types():
    """Only labor invoices present → 'Labor invoices' sheet created, Released Funds/Materials skipped."""
    ctx = _make_context()
    labor_inv = _make_invoice(invoice_type=InvoiceType.LABOR, amount=Decimal("200.00"))
    bundle = _make_bundle([labor_inv])
    wb = openpyxl.load_workbook(BytesIO(build_xlsx(ctx, bundle)))
    assert "Labor invoices" in wb.sheetnames
    assert "Released Funds invoices" not in wb.sheetnames
    assert "Materials & Services invoices" not in wb.sheetnames


def test_currency_cells_are_floats_with_eur_format():
    """Currency cells in Summary carry EUR_FR_FORMAT number_format and numeric values.

    Note: openpyxl reads back whole-number floats (e.g. 200.0) as int after a
    write-read roundtrip when the value has no fractional part. We therefore
    accept both int and float — what matters is (a) the number_format is set
    correctly and (b) the value is numeric (not a string).
    """
    ctx = _make_context()
    # Use a fractional amount to force a true float on roundtrip
    inv = _make_invoice(amount=Decimal("200.50"))
    bundle = _make_bundle([inv])
    wb = openpyxl.load_workbook(BytesIO(build_xlsx(ctx, bundle)))
    ws = wb["Summary"]

    # Scan all cells carrying EUR_FR_FORMAT
    eur_cells = [
        cell
        for row in ws.iter_rows()
        for cell in row
        if cell.number_format == EUR_FR_FORMAT and isinstance(cell.value, (int, float))
    ]
    assert len(eur_cells) >= 1, "Expected at least one EUR-formatted numeric cell in Summary"
    # Verify values are numeric (not strings)
    for cell in eur_cells:
        assert isinstance(cell.value, (int, float)), f"Expected numeric, got {type(cell.value)}: {cell.value}"


def test_grand_total_row_present():
    """Summary sheet has a 'GRAND TOTAL' row in the invoices section."""
    ctx = _make_context()
    invoices = [
        _make_invoice(invoice_number="INV-001", amount=Decimal("100.00")),
        _make_invoice(invoice_number="INV-002", amount=Decimal("200.00")),
    ]
    bundle = _make_bundle(invoices)
    wb = openpyxl.load_workbook(BytesIO(build_xlsx(ctx, bundle)))
    ws = wb["Summary"]

    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "GRAND TOTAL" in all_values, "Expected 'GRAND TOTAL' text in Summary sheet"


def test_empty_range_only_summary_sheet():
    """Empty bundle → only 'Summary' sheet, no type sheets."""
    ctx = _make_context()
    bundle = _make_bundle(invoices=[])
    wb = openpyxl.load_workbook(BytesIO(build_xlsx(ctx, bundle)))
    assert wb.sheetnames == ["Summary"], f"Expected only Summary sheet, got {wb.sheetnames}"


def test_xss_safe_recipient_name():
    """Script-tag recipient name must not crash openpyxl XML serialization."""
    ctx = _make_context()
    xss_inv = _make_invoice(recipient="<script>alert(1)</script>", amount=Decimal("50.00"))
    bundle = _make_bundle([xss_inv])

    # Must not raise — openpyxl escapes XML entities at write time
    raw = build_xlsx(ctx, bundle)
    assert raw[:4] == b"PK\x03\x04", "Expected valid xlsx magic bytes after XSS recipient name"

    # Verify workbook is loadable (XML is well-formed)
    wb = openpyxl.load_workbook(BytesIO(raw))
    assert "Summary" in wb.sheetnames
