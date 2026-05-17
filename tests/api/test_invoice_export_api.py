"""API smoke tests for GET /api/v1/projects/<id>/invoices-export.

Mirrors tests/api/test_labor_export_endpoint.py fixture style.

Covers:
- 401 — no auth token
- 422 — missing format param
- 422 — invalid format value
- 422 — range exceeds 24 months
- 422 — from > to
- 404 — unknown project
- 200 xlsx smoke — magic bytes PK, content-type, Content-Disposition
- 200 pdf smoke — magic bytes %PDF-, content-type
- type filter in filename — ?type=labor → filename has '-labor'
- multi-invoice xlsx + pdf smoke (MED-7)
"""

from __future__ import annotations

import datetime
from io import BytesIO
from uuid import uuid4

import pytest

from app.infrastructure.database.models import (
    InvoiceModel,
    PermissionModel,
    ProjectModel,
    RoleModel,
    UserModel,
)


# ---------------------------------------------------------------------------
# App fixture
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def inv_export_app():
    """Flask app with in-memory DB + full invoice + export container for route tests."""
    from app import create_app, db
    from app.infrastructure.adapters.argon2_hasher import Argon2PasswordHasher
    from app.infrastructure.adapters.flask_session import FlaskSessionManager
    from app.infrastructure.adapters.jwt_issuer import JWTTokenIssuer
    from app.infrastructure.adapters.sqlalchemy_invoice import SQLAlchemyInvoiceRepository
    from app.infrastructure.adapters.sqlalchemy_project import SQLAlchemyProjectRepository
    from app.infrastructure.adapters.sqlalchemy_user import SQLAlchemyUserRepository
    from config import TestingConfig
    from wiring import configure_container

    class InvExportTestConfig(TestingConfig):
        JWT_TOKEN_LOCATION = ["headers", "cookies"]
        RATELIMIT_ENABLED = False
        RATELIMIT_STORAGE_URI = "memory://"

    test_app = create_app(InvExportTestConfig)

    with test_app.app_context():
        db.create_all()

        hasher = Argon2PasswordHasher()
        user_repo = SQLAlchemyUserRepository(db.session)
        project_repo = SQLAlchemyProjectRepository(db.session)
        invoice_repo = SQLAlchemyInvoiceRepository(db.session)

        # Permissions
        read_perm = PermissionModel(name="project:read", resource="project", action="read")

        # Role with project:read only — *:* omitted so require_permission("project:read") is exercised
        admin_role = RoleModel(name="inv_export_admin", description="Inv Export Admin")
        admin_role.permissions.append(read_perm)

        # Role without project:read
        noperm_role = RoleModel(name="inv_export_noperm", description="No Read")

        db.session.add_all([read_perm, admin_role, noperm_role])
        db.session.flush()

        # Admin user
        admin_user = UserModel(
            email="invexportadmin@test.com",
            password_hash=hasher.hash("Admin1234!"),
            is_active=True,
        )
        admin_user.roles.append(admin_role)
        db.session.add(admin_user)
        db.session.flush()

        # Unprivileged user
        noperm_user = UserModel(
            email="invexportnoperm@test.com",
            password_hash=hasher.hash("Admin1234!"),
            is_active=True,
        )
        noperm_user.roles.append(noperm_role)
        db.session.add(noperm_user)
        db.session.flush()

        # Project
        project = ProjectModel(
            name="Invoice Export Test Project",
            owner_id=admin_user.id,
        )
        db.session.add(project)
        db.session.commit()

        configure_container(
            user_repository=user_repo,
            project_repository=project_repo,
            password_hasher=hasher,
            token_issuer=JWTTokenIssuer(),
            session_manager=FlaskSessionManager(),
            invoice_repository=invoice_repo,
        )

        test_app._test_admin_email = "invexportadmin@test.com"
        test_app._test_admin_password = "Admin1234!"
        test_app._test_project_id = str(project.id)

        yield test_app

        db.session.remove()
        db.drop_all()


@pytest.fixture
def inv_export_client(inv_export_app):
    return inv_export_app.test_client()


def _login(client, email: str, password: str) -> str:
    resp = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert resp.status_code == 200, f"Login failed: {resp.get_data(as_text=True)}"
    return resp.get_json()["access_token"]


@pytest.fixture
def admin_token(inv_export_client, inv_export_app):
    return _login(inv_export_client, inv_export_app._test_admin_email, inv_export_app._test_admin_password)


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _export_url(project_id: str) -> str:
    return f"/api/v1/projects/{project_id}/invoices-export"


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_unauth_returns_401(inv_export_client, inv_export_app):
    """No Authorization header → 401."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01", "format": "xlsx"},
    )
    assert resp.status_code == 401


def test_missing_format_returns_422(inv_export_client, inv_export_app, admin_token):
    """Missing 'format' param → 422 validation_error."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 422
    assert resp.get_json()["error"] == "validation_error"


def test_invalid_format_returns_422(inv_export_client, inv_export_app, admin_token):
    """format=csv → 422 validation_error."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01", "format": "csv"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 422
    data = resp.get_json()
    assert data["error"] == "validation_error"


def test_range_too_large_returns_422(inv_export_client, inv_export_app, admin_token):
    """25-month range (from=2024-01, to=2026-02) → 422."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2024-01", "to": "2026-02", "format": "xlsx"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 422
    assert resp.get_json()["error"] == "validation_error"


def test_from_after_to_returns_422(inv_export_client, inv_export_app, admin_token):
    """from > to → 422 validation_error."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-06", "to": "2026-01", "format": "xlsx"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 422
    assert resp.get_json()["error"] == "validation_error"


def test_unknown_project_returns_404(inv_export_client, inv_export_app, admin_token):
    """Non-existent project UUID → 404.

    @require_project_access() intercepts before route body and returns {"error": "NotFound"}.
    """
    url = _export_url(str(uuid4()))
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01", "format": "xlsx"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 404
    data = resp.get_json()
    assert data["error"] in ("NotFound", "project_not_found")


def test_xlsx_smoke(inv_export_client, inv_export_app, admin_token):
    """200 xlsx: PK magic bytes, correct content-type, attachment Content-Disposition."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01", "format": "xlsx"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 200, resp.get_data(as_text=True)
    assert resp.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.data[:4] == b"PK\x03\x04", "Expected xlsx ZIP magic bytes"

    cd = resp.headers.get("Content-Disposition", "")
    assert "attachment" in cd
    assert "invoices-" in cd
    assert ".xlsx" in cd


def test_pdf_smoke(inv_export_client, inv_export_app, admin_token):
    """200 pdf: %PDF- magic bytes, correct content-type."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01", "format": "pdf"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 200, resp.get_data(as_text=True)
    assert resp.content_type == "application/pdf"
    assert resp.data[:5] == b"%PDF-", "Expected PDF magic bytes"


def test_type_filter_in_filename(inv_export_client, inv_export_app, admin_token):
    """?type=labor → Content-Disposition filename contains '-labor'."""
    url = _export_url(inv_export_app._test_project_id)
    resp = inv_export_client.get(
        url,
        query_string={"from": "2026-01", "to": "2026-01", "format": "xlsx", "type": "labor"},
        headers=_auth(admin_token),
    )
    assert resp.status_code == 200, resp.get_data(as_text=True)
    cd = resp.headers.get("Content-Disposition", "")
    assert "-labor" in cd, f"Expected '-labor' suffix in Content-Disposition; got: {cd}"
    assert ".xlsx" in cd


# ---------------------------------------------------------------------------
# MED-7 — multi-invoice xlsx + pdf integration smoke
# ---------------------------------------------------------------------------


def test_multi_invoice_xlsx_and_pdf_smoke(inv_export_client, inv_export_app, admin_token):
    """Seed 3 invoices (2 client + 1 labor), export xlsx and pdf, validate content.

    xlsx assertions:
    - 200, xlsx content-type, PK magic bytes
    - openpyxl opens successfully
    - Sheet names include Summary, Released Funds invoices, Labor invoices
    - "Materials & Services invoices" sheet absent (no materials & services invoices seeded)
    - Summary sheet contains GRAND TOTAL label

    pdf assertions:
    - 200, pdf content-type, %PDF- magic bytes
    - At least 4 /Type /Page markers (1 summary + 3 invoices)
    """
    import openpyxl

    from app import db

    project_id_str = inv_export_app._test_project_id

    # --- Seed 3 invoices directly via ORM ---
    with inv_export_app.app_context():
        import uuid

        project_uuid = uuid.UUID(project_id_str)
        items_payload = [{"description": "Service", "quantity": 1.0, "unit_price": 100.0}]

        inv1 = InvoiceModel(
            id=uuid4(),
            project_id=project_uuid,
            invoice_number="MULTI-C001",
            type="released_funds",
            issue_date=datetime.date(2026, 3, 1),
            recipient_name="Client Alpha",
            items=items_payload,
        )
        inv2 = InvoiceModel(
            id=uuid4(),
            project_id=project_uuid,
            invoice_number="MULTI-C002",
            type="released_funds",
            issue_date=datetime.date(2026, 3, 15),
            recipient_name="Client Beta",
            items=items_payload,
        )
        inv3 = InvoiceModel(
            id=uuid4(),
            project_id=project_uuid,
            invoice_number="MULTI-L001",
            type="labor",
            issue_date=datetime.date(2026, 3, 20),
            recipient_name="Labor Gamma",
            items=items_payload,
        )
        db.session.add_all([inv1, inv2, inv3])
        db.session.commit()

    url = _export_url(project_id_str)

    # --- xlsx ---
    resp_xlsx = inv_export_client.get(
        url,
        query_string={"from": "2026-03", "to": "2026-03", "format": "xlsx"},
        headers=_auth(admin_token),
    )
    assert resp_xlsx.status_code == 200, resp_xlsx.get_data(as_text=True)
    assert resp_xlsx.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp_xlsx.data[:4] == b"PK\x03\x04", "Expected xlsx ZIP magic bytes"

    wb = openpyxl.load_workbook(BytesIO(resp_xlsx.data), data_only=True)
    sheet_names = wb.sheetnames

    assert "Summary" in sheet_names, f"Missing 'Summary' sheet; got: {sheet_names}"
    assert "Released Funds invoices" in sheet_names, f"Missing 'Released Funds invoices' sheet; got: {sheet_names}"
    assert "Labor invoices" in sheet_names, f"Missing 'Labor invoices' sheet; got: {sheet_names}"
    assert (
        "Materials & Services invoices" not in sheet_names
    ), f"Unexpected 'Materials & Services invoices' sheet (none seeded); got: {sheet_names}"

    # Summary sheet must contain GRAND TOTAL label somewhere
    ws_summary = wb["Summary"]
    all_summary_values = [ws_summary.cell(row=r, column=1).value for r in range(1, 30)]
    has_grand_total = any(v and "GRAND TOTAL" in str(v).upper() for v in all_summary_values)
    assert has_grand_total, f"GRAND TOTAL label not found in Summary col A: {all_summary_values}"

    # --- pdf ---
    resp_pdf = inv_export_client.get(
        url,
        query_string={"from": "2026-03", "to": "2026-03", "format": "pdf"},
        headers=_auth(admin_token),
    )
    assert resp_pdf.status_code == 200, resp_pdf.get_data(as_text=True)
    assert resp_pdf.content_type == "application/pdf"
    assert resp_pdf.data[:5] == b"%PDF-", "Expected PDF magic bytes"

    page_count = resp_pdf.data.count(b"/Type /Page")
    assert page_count >= 4, f"Expected >= 4 /Type /Page markers (1 summary + 3 invoices); got {page_count}"
